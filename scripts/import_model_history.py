#!/usr/bin/env python3
"""
import_model_history.py — Import all IWP model tabs from xlsx and compute performance.

Usage:
    python scripts/import_model_history.py [--xlsx PATH] [--skip-prices]
"""

import os
import re
import logging
import argparse
import warnings
from datetime import datetime, date, timedelta
from pathlib import Path
from collections import defaultdict

warnings.filterwarnings("ignore")

import psycopg2
import psycopg2.extras
import openpyxl
import yfinance as yf

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

PROJECT_DIR = Path(__file__).parent.parent

def load_env():
    for line in (PROJECT_DIR / ".env").read_text().splitlines():
        if "=" in line and not line.startswith("#"):
            k, v = line.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip().strip('"'))

load_env()
DB_URL = os.environ["DATABASE_URL"]

DEFAULT_XLSX = "/Users/nikorrw/Downloads/IWP Models Historic Allocations.xlsx"
BENCHMARKS   = ["SPY", "BND"]

# Map sheet name → stable portfolio_id slug
def sheet_to_id(name):
    return re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")

def get_db():
    return psycopg2.connect(DB_URL, cursor_factory=psycopg2.extras.RealDictCursor)


# ─── Migration ────────────────────────────────────────────────────────────────

def run_migration(db):
    sql = (PROJECT_DIR / "database/migrations/008_models.sql").read_text()
    try:
        with db.cursor() as cur:
            cur.execute(sql)
        db.commit()
        log.info("Migration OK")
    except Exception as e:
        db.rollback()
        if "already exists" in str(e):
            log.info("Tables already exist")
        else:
            raise


# ─── xlsx import ─────────────────────────────────────────────────────────────

def import_xlsx(db, xlsx_path):
    """Import all sheets. Returns {portfolio_id: {date: [holdings]}}."""
    log.info("Opening %s", xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    all_snapshots = {}   # portfolio_id → {date → [holdings]}
    total_rows = 0

    for sheet_name in wb.sheetnames:
        ws   = wb[sheet_name]
        pid  = sheet_to_id(sheet_name)
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Find header row
        header_idx = 0
        for i, row in enumerate(rows):
            if row and str(row[0]).strip().lower() == "ticker":
                header_idx = i
                break

        snapshots = defaultdict(list)
        for row in rows[header_idx + 1:]:
            if not row or row[0] is None:
                continue
            ticker = str(row[0]).strip().upper()
            name   = str(row[1]).strip() if row[1] else ""
            # Clean markdown artifacts
            name = re.sub(r"\*\*", "", name)
            if "can be simplified to:" in name:
                name = name.split("can be simplified to:")[0].strip().rstrip(".")
            try:
                weight = round(float(row[2]), 6)
            except (TypeError, ValueError):
                continue
            raw_date = row[3]
            if isinstance(raw_date, datetime):
                snap_date = raw_date.date()
            elif isinstance(raw_date, date):
                snap_date = raw_date
            elif isinstance(raw_date, str):
                for fmt in ("%m/%d/%Y", "%Y-%m-%d"):
                    try:
                        snap_date = datetime.strptime(raw_date.strip(), fmt).date()
                        break
                    except ValueError:
                        pass
                else:
                    continue
            else:
                continue

            snapshots[snap_date].append({"ticker": ticker, "name": name, "weight": weight})

        # Seed portfolio record
        inception = min(snapshots.keys()) if snapshots else None
        with db.cursor() as cur:
            cur.execute("""
                INSERT INTO model_portfolios (id, name, description, inception_date)
                VALUES (%s, %s, %s, %s)
                ON CONFLICT (id) DO UPDATE SET name=EXCLUDED.name, inception_date=EXCLUDED.inception_date
            """, (pid, sheet_name, f"IWP {sheet_name} model", inception))

        # Bulk upsert allocations
        rows_upserted = 0
        with db.cursor() as cur:
            for snap_date, holdings in snapshots.items():
                for h in holdings:
                    cur.execute("""
                        INSERT INTO model_allocations
                            (portfolio_id, snapshot_date, ticker, instrument_name, weight)
                        VALUES (%s, %s, %s, %s, %s)
                        ON CONFLICT (portfolio_id, snapshot_date, ticker) DO UPDATE
                            SET instrument_name=EXCLUDED.instrument_name, weight=EXCLUDED.weight
                    """, (pid, snap_date, h["ticker"], h["name"], h["weight"]))
                    rows_upserted += 1

        db.commit()
        n_snap  = len(snapshots)
        n_tick  = len({h["ticker"] for hlist in snapshots.values() for h in hlist})
        total_rows += rows_upserted
        log.info("  %-25s → %s | %3d snapshots | %3d tickers | %4d rows",
                 sheet_name, pid, n_snap, n_tick, rows_upserted)
        all_snapshots[pid] = snapshots

    log.info("Total rows imported: %d across %d portfolios", total_rows, len(wb.sheetnames))
    return all_snapshots


# ─── Price fetching ───────────────────────────────────────────────────────────

def fetch_all_prices(db, all_snapshots):
    all_tickers = set(BENCHMARKS)
    all_dates   = []
    for snapshots in all_snapshots.values():
        all_dates.extend(snapshots.keys())
        for holdings in snapshots.values():
            for h in holdings:
                if h["ticker"] != "CASH":
                    all_tickers.add(h["ticker"])

    if not all_dates:
        return

    start = min(all_dates) - timedelta(days=14)
    end   = min(date.today(), max(all_dates) + timedelta(days=7))

    log.info("Bulk downloading %d tickers (%s → %s)…", len(all_tickers), start, end)

    ticker_list = sorted(all_tickers)
    try:
        data = yf.download(
            ticker_list,
            start=start.isoformat(),
            end=end.isoformat(),
            auto_adjust=True,
            progress=False,
        )
        if data.empty:
            log.warning("yfinance returned no data")
            return

        close = data["Close"] if "Close" in data.columns else data
        to_insert = []
        for col in close.columns:
            ticker = str(col).upper()
            series = close[col].dropna()
            for idx, price_val in series.items():
                pdate = idx.date() if hasattr(idx, "date") else idx
                to_insert.append((ticker, pdate, float(price_val)))

        with db.cursor() as cur:
            psycopg2.extras.execute_values(cur, """
                INSERT INTO price_cache (ticker, price_date, close_price)
                VALUES %s ON CONFLICT DO NOTHING
            """, to_insert, page_size=1000)
        db.commit()
        log.info("Cached %d price rows", len(to_insert))
    except Exception as e:
        log.error("Bulk download failed: %s", e)


def get_price(db, ticker, target_date):
    if ticker == "CASH":
        return 1.0
    with db.cursor() as cur:
        cur.execute("""
            SELECT close_price FROM price_cache
            WHERE ticker = %s
              AND price_date BETWEEN %s AND %s
            ORDER BY ABS(price_date - %s::date)
            LIMIT 1
        """, (ticker, target_date - timedelta(days=7), target_date + timedelta(days=3), target_date))
        row = cur.fetchone()
        return float(row["close_price"]) if row else None


# ─── Performance computation ──────────────────────────────────────────────────

def compute_performance(db, portfolio_id):
    with db.cursor() as cur:
        cur.execute("""
            SELECT DISTINCT snapshot_date FROM model_allocations
            WHERE portfolio_id = %s ORDER BY snapshot_date
        """, (portfolio_id,))
        dates = [r["snapshot_date"] for r in cur.fetchall()]

    if not dates:
        return

    today         = date.today()
    cum_port      = 1.0
    cum_spy       = 1.0
    cum_blend     = 1.0
    results       = []

    for i, snap_date in enumerate(dates):
        next_date = dates[i + 1] if i + 1 < len(dates) else today

        with db.cursor() as cur:
            cur.execute("SELECT ticker, weight FROM model_allocations WHERE portfolio_id=%s AND snapshot_date=%s",
                        (portfolio_id, snap_date))
            holdings = cur.fetchall()

        period_ret     = 0.0
        weight_covered = 0.0
        pos_rows       = []

        for h in holdings:
            ticker = h["ticker"]
            weight = float(h["weight"])
            p0     = get_price(db, ticker, snap_date)
            p1     = get_price(db, ticker, next_date)
            if ticker == "CASH":
                ret = 0.0; p0 = p1 = 1.0
            elif p0 and p1 and p0 > 0:
                ret = (p1 - p0) / p0
            else:
                ret = None
            if ret is not None:
                period_ret     += weight * ret
                weight_covered += weight
            pos_rows.append({"ticker": ticker, "weight": weight,
                             "p0": p0, "p1": p1, "ret": ret,
                             "contrib": weight * ret if ret is not None else None})

        if 0 < weight_covered < 0.99:
            period_ret = period_ret / weight_covered * min(weight_covered + 0.05, 1.0)

        spy0   = get_price(db, "SPY", snap_date)
        spy1   = get_price(db, "SPY", next_date)
        bnd0   = get_price(db, "BND", snap_date)
        bnd1   = get_price(db, "BND", next_date)
        spy_r  = (spy1 - spy0) / spy0 if spy0 and spy1 else None
        bnd_r  = (bnd1 - bnd0) / bnd0 if bnd0 and bnd1 else None
        b60_r  = 0.60 * spy_r + 0.40 * bnd_r if spy_r is not None and bnd_r is not None else None

        cum_port  *= (1 + period_ret)
        if spy_r  is not None: cum_spy   *= (1 + spy_r)
        if b60_r  is not None: cum_blend *= (1 + b60_r)

        with db.cursor() as cur:
            cur.execute("""
                INSERT INTO model_snapshot_performance (
                    portfolio_id, snapshot_date, next_snapshot_date, period_days,
                    period_return, spy_return, agg_return, blend_6040_return,
                    excess_vs_spy, excess_vs_6040,
                    cumulative_return, spy_cumulative, blend_cumulative,
                    tickers_priced, tickers_total, weight_covered
                ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (portfolio_id, snapshot_date) DO UPDATE SET
                    next_snapshot_date=EXCLUDED.next_snapshot_date,
                    period_days=EXCLUDED.period_days,
                    period_return=EXCLUDED.period_return,
                    spy_return=EXCLUDED.spy_return, agg_return=EXCLUDED.agg_return,
                    blend_6040_return=EXCLUDED.blend_6040_return,
                    excess_vs_spy=EXCLUDED.excess_vs_spy, excess_vs_6040=EXCLUDED.excess_vs_6040,
                    cumulative_return=EXCLUDED.cumulative_return,
                    spy_cumulative=EXCLUDED.spy_cumulative, blend_cumulative=EXCLUDED.blend_cumulative,
                    tickers_priced=EXCLUDED.tickers_priced, tickers_total=EXCLUDED.tickers_total,
                    weight_covered=EXCLUDED.weight_covered, computed_at=NOW()
            """, (
                portfolio_id, snap_date, next_date, (next_date - snap_date).days,
                round(period_ret, 6),
                round(spy_r,  6) if spy_r  is not None else None,
                round(bnd_r,  6) if bnd_r  is not None else None,
                round(b60_r,  6) if b60_r  is not None else None,
                round(period_ret - spy_r,  6) if spy_r  is not None else None,
                round(period_ret - b60_r,  6) if b60_r  is not None else None,
                round(cum_port  - 1, 6),
                round(cum_spy   - 1, 6),
                round(cum_blend - 1, 6),
                sum(1 for p in pos_rows if p["ret"] is not None),
                len(pos_rows),
                round(weight_covered, 4),
            ))
            for p in pos_rows:
                cur.execute("""
                    INSERT INTO model_position_returns
                        (portfolio_id, snapshot_date, ticker, weight,
                         price_start, price_end, position_return, contribution)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                    ON CONFLICT (portfolio_id, snapshot_date, ticker) DO UPDATE SET
                        price_start=EXCLUDED.price_start, price_end=EXCLUDED.price_end,
                        position_return=EXCLUDED.position_return, contribution=EXCLUDED.contribution
                """, (portfolio_id, snap_date, p["ticker"], p["weight"],
                      p["p0"], p["p1"],
                      round(p["ret"],    6) if p["ret"]    is not None else None,
                      round(p["contrib"],6) if p["contrib"] is not None else None))
        db.commit()

    log.info("  %-25s cumul=%+.1f%%  spy=%+.1f%%  60/40=%+.1f%%",
             portfolio_id, (cum_port-1)*100, (cum_spy-1)*100, (cum_blend-1)*100)


# ─── Summary ──────────────────────────────────────────────────────────────────

def print_summary(db):
    with db.cursor() as cur:
        cur.execute("""
            SELECT mp.name,
                   COUNT(DISTINCT ma.snapshot_date) AS snapshots,
                   MIN(ma.snapshot_date) AS from_date,
                   MAX(ma.snapshot_date) AS to_date,
                   (SELECT cumulative_return FROM model_snapshot_performance msp
                    WHERE msp.portfolio_id = mp.id ORDER BY snapshot_date DESC LIMIT 1) AS cum_ret,
                   (SELECT spy_cumulative FROM model_snapshot_performance msp
                    WHERE msp.portfolio_id = mp.id ORDER BY snapshot_date DESC LIMIT 1) AS spy_ret,
                   (SELECT blend_cumulative FROM model_snapshot_performance msp
                    WHERE msp.portfolio_id = mp.id ORDER BY snapshot_date DESC LIMIT 1) AS blend_ret
            FROM model_portfolios mp
            JOIN model_allocations ma ON ma.portfolio_id = mp.id
            GROUP BY mp.id, mp.name
            ORDER BY mp.name
        """)
        rows = cur.fetchall()

    print("\n" + "═" * 85)
    print(f"  {'MODEL':<26} {'FROM':>10}  {'TO':>10}  {'SNAPS':>5}  {'IWP':>8}  {'SPY':>8}  {'60/40':>8}  {'α SPY':>8}")
    print("─" * 85)
    for r in rows:
        cr = float(r["cum_ret"]   or 0)
        sr = float(r["spy_ret"]   or 0)
        br = float(r["blend_ret"] or 0)
        print(f"  {r['name']:<26} {str(r['from_date']):>10}  {str(r['to_date']):>10}  "
              f"{r['snapshots']:>5}  {cr*100:>+7.1f}%  {sr*100:>+7.1f}%  {br*100:>+7.1f}%  {(cr-sr)*100:>+7.1f}%")
    print("═" * 85)


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx",         default=DEFAULT_XLSX)
    parser.add_argument("--skip-prices",  action="store_true")
    parser.add_argument("--skip-import",  action="store_true")
    args = parser.parse_args()

    db = get_db()
    run_migration(db)

    if not args.skip_import:
        all_snapshots = import_xlsx(db, args.xlsx)
    else:
        # rebuild snapshot dict from DB for price fetching
        all_snapshots = {}
        with db.cursor() as cur:
            cur.execute("SELECT id FROM model_portfolios")
            for pid_row in cur.fetchall():
                pid = pid_row["id"]
                cur.execute("SELECT DISTINCT snapshot_date FROM model_allocations WHERE portfolio_id=%s", (pid,))
                dates = [r["snapshot_date"] for r in cur.fetchall()]
                snaps = {}
                for d in dates:
                    cur.execute("SELECT ticker, weight FROM model_allocations WHERE portfolio_id=%s AND snapshot_date=%s",
                                (pid, d))
                    snaps[d] = [dict(r) for r in cur.fetchall()]
                all_snapshots[pid] = snaps

    if not args.skip_prices:
        fetch_all_prices(db, all_snapshots)

    log.info("Computing performance for all portfolios…")
    with db.cursor() as cur:
        cur.execute("SELECT id, name FROM model_portfolios ORDER BY name")
        portfolios = cur.fetchall()

    for p in portfolios:
        log.info("  → %s", p["name"])
        compute_performance(db, p["id"])

    print_summary(db)
    db.close()


if __name__ == "__main__":
    main()
